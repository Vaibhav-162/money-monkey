"""One master IPO sheet with grouped headers (group row + subcolumn row)."""

from __future__ import annotations

import os
import re
import time
from pathlib import Path
from typing import Any, Iterable, Optional

import pandas as pd

# (group title, field key, subcolumn label)
COLUMN_GROUPS: list[tuple[str, str, str]] = [
    ("Identity", "ipo_id", "IPO ID"),
    ("Identity", "company_name", "Company"),
    ("Identity", "slug", "Slug"),
    ("Identity", "url", "URL"),
    ("Identity", "exchange_type", "Board"),
    ("Identity", "listing_year", "Listing year"),
    ("Classification", "industry", "Industry"),
    ("Classification", "issue_type", "Issue type"),
    ("Classification", "sale_type", "Sale type"),
    ("Classification", "listing_at", "Listing at"),
    ("Calendar", "ipo_open", "Open"),
    ("Calendar", "ipo_close", "Close"),
    ("Calendar", "allotment_date", "Allotment"),
    ("Calendar", "refund_date", "Refund"),
    ("Calendar", "credit_date", "Credit"),
    ("Calendar", "listing_date", "Listing"),
    ("Calendar", "anchor_bid_date", "Anchor bid"),
    ("Pricing", "face_value", "Face value"),
    ("Pricing", "price_band_low", "Band low"),
    ("Pricing", "price_band_high", "Band high"),
    ("Pricing", "issue_price", "Issue price"),
    ("Pricing", "lot_size", "Lot size"),
    ("Pricing", "retail_min_amount", "Retail min amt"),
    ("Issue size", "issue_size_shares", "Total shares"),
    ("Issue size", "issue_size_cr", "Total Cr"),
    ("Issue size", "fresh_issue_shares", "Fresh shares"),
    ("Issue size", "fresh_issue_cr", "Fresh Cr"),
    ("Issue size", "ofs_shares", "OFS shares"),
    ("Issue size", "ofs_cr", "OFS Cr"),
    ("Issue size", "pre_issue_shares", "Pre-issue shares"),
    ("Issue size", "post_issue_shares", "Post-issue shares"),
    ("Identifiers", "bse_code", "BSE code"),
    ("Identifiers", "nse_symbol", "NSE symbol"),
    ("Identifiers", "isin", "ISIN"),
    ("Identifiers", "market_cap_listing_cr", "Mcap listing Cr"),
    ("Reservation %", "qib_pct", "QIB"),
    ("Reservation %", "anchor_pct", "Anchor"),
    ("Reservation %", "nii_pct", "NII"),
    ("Reservation %", "retail_pct", "Retail"),
    ("Reservation %", "employee_pct", "Employee"),
    ("Reservation %", "market_maker_pct", "Market maker"),
    ("Anchor", "anchor_shares", "Shares"),
    ("Anchor", "anchor_amount_cr", "Amount Cr"),
    ("Anchor", "anchor_lockin_30d", "Lock-in 30d"),
    ("Anchor", "anchor_lockin_90d", "Lock-in 90d"),
    ("Lot size", "retail_min_lots", "Retail min lots"),
    ("Lot size", "retail_min_shares", "Retail min shares"),
    ("Lot size", "retail_max_lots", "Retail max lots"),
    ("Lot size", "retail_max_shares", "Retail max shares"),
    ("Lot size", "retail_max_amount", "Retail max amt"),
    ("Lot size", "shni_min_lots", "sHNI min lots"),
    ("Lot size", "shni_min_shares", "sHNI min shares"),
    ("Lot size", "shni_min_amount", "sHNI min amt"),
    ("Lot size", "shni_max_lots", "sHNI max lots"),
    ("Lot size", "shni_max_shares", "sHNI max shares"),
    ("Lot size", "shni_max_amount", "sHNI max amt"),
    ("Lot size", "bhni_min_lots", "bHNI min lots"),
    ("Lot size", "bhni_min_shares", "bHNI min shares"),
    ("Lot size", "bhni_min_amount", "bHNI min amt"),
    ("Financials FY1 (latest)", "fy1_period", "Period"),
    ("Financials FY1 (latest)", "fy1_assets", "Assets Cr"),
    ("Financials FY1 (latest)", "fy1_total_income", "Total income Cr"),
    ("Financials FY1 (latest)", "fy1_pat", "PAT Cr"),
    ("Financials FY1 (latest)", "fy1_ebitda", "EBITDA Cr"),
    ("Financials FY1 (latest)", "fy1_net_worth", "Net worth Cr"),
    ("Financials FY1 (latest)", "fy1_borrowings", "Borrowings Cr"),
    ("Financials FY2", "fy2_period", "Period"),
    ("Financials FY2", "fy2_assets", "Assets Cr"),
    ("Financials FY2", "fy2_total_income", "Total income Cr"),
    ("Financials FY2", "fy2_pat", "PAT Cr"),
    ("Financials FY2", "fy2_ebitda", "EBITDA Cr"),
    ("Financials FY2", "fy2_net_worth", "Net worth Cr"),
    ("Financials FY2", "fy2_borrowings", "Borrowings Cr"),
    ("Financials FY3", "fy3_period", "Period"),
    ("Financials FY3", "fy3_assets", "Assets Cr"),
    ("Financials FY3", "fy3_total_income", "Total income Cr"),
    ("Financials FY3", "fy3_pat", "PAT Cr"),
    ("Financials FY3", "fy3_ebitda", "EBITDA Cr"),
    ("Financials FY3", "fy3_net_worth", "Net worth Cr"),
    ("Financials FY3", "fy3_borrowings", "Borrowings Cr"),
    ("Valuation", "eps_pre", "EPS pre"),
    ("Valuation", "eps_post", "EPS post"),
    ("Valuation", "pe_pre", "P/E pre"),
    ("Valuation", "pe_post", "P/E post"),
    ("KPI (latest)", "roe", "ROE"),
    ("KPI (latest)", "roce", "ROCE"),
    ("KPI (latest)", "ronw", "RoNW"),
    ("KPI (latest)", "debt_equity", "D/E"),
    ("KPI (latest)", "pat_margin", "PAT margin"),
    ("KPI (latest)", "ebitda_margin", "EBITDA margin"),
    ("KPI (latest)", "nav", "NAV"),
    ("KPI (latest)", "pbv", "P/B"),
    ("Shareholding", "promoter_pre_pct", "Promoter pre %"),
    ("Shareholding", "promoter_post_pct", "Promoter post %"),
    ("Shareholding", "public_pre_pct", "Public pre %"),
    ("Shareholding", "public_post_pct", "Public post %"),
    ("Shareholding", "promoters", "Promoters"),
    ("GMP at close", "gmp_close_date", "Date"),
    ("GMP at close", "gmp_rs", "GMP Rs"),
    ("GMP at close", "gmp_pct", "GMP %"),
    ("GMP at close", "gmp_est_listing_price", "Est. listing"),
    ("GMP at close", "kostak_rs", "Kostak"),
    ("GMP at close", "subject_to_sauda", "Subject to sauda"),
    ("GMP at close", "gmp_obs_count", "History rows"),
    ("Subscription", "sub_qib_x", "QIB x"),
    ("Subscription", "sub_nii_x", "NII x"),
    ("Subscription", "sub_bnii_x", "bNII x"),
    ("Subscription", "sub_snii_x", "sNII x"),
    ("Subscription", "sub_retail_x", "Retail x"),
    ("Subscription", "sub_total_x", "Total x"),
    ("Subscription", "total_applications", "Applications"),
    ("Reviews", "broker_subscribe", "Subscribe"),
    ("Reviews", "broker_may_apply", "May apply"),
    ("Reviews", "broker_neutral", "Neutral"),
    ("Reviews", "broker_avoid", "Avoid"),
    ("Intermediaries", "registrar", "Registrar"),
    ("Intermediaries", "lead_managers", "Lead managers"),
    ("Tracker performance", "listing_day_close", "Listing close"),
    ("Tracker performance", "listing_day_gain_pct", "Listing gain %"),
    ("Tracker performance", "current_price", "Current price"),
    ("Tracker performance", "profit_loss_pct", "P/L %"),
    ("Listing day BSE", "listing_bse_open", "Open"),
    ("Listing day BSE", "listing_bse_high", "High"),
    ("Listing day BSE", "listing_bse_low", "Low"),
    ("Listing day BSE", "listing_bse_last", "Last"),
    ("Listing day NSE", "listing_nse_open", "Open"),
    ("Listing day NSE", "listing_nse_high", "High"),
    ("Listing day NSE", "listing_nse_low", "Low"),
    ("Listing day NSE", "listing_nse_last", "Last"),
    ("Objects of issue", "object_1", "Object 1"),
    ("Objects of issue", "object_1_cr", "Amt 1 Cr"),
    ("Objects of issue", "object_2", "Object 2"),
    ("Objects of issue", "object_2_cr", "Amt 2 Cr"),
    ("Objects of issue", "object_3", "Object 3"),
    ("Objects of issue", "object_3_cr", "Amt 3 Cr"),
    ("Objects of issue", "object_4", "Object 4"),
    ("Objects of issue", "object_4_cr", "Amt 4 Cr"),
    ("OFS sellers", "ofs_1_name", "Seller 1"),
    ("OFS sellers", "ofs_1_category", "Cat 1"),
    ("OFS sellers", "ofs_1_shares", "Shares 1"),
    ("OFS sellers", "ofs_1_cr", "Amt 1 Cr"),
    ("OFS sellers", "ofs_2_name", "Seller 2"),
    ("OFS sellers", "ofs_2_category", "Cat 2"),
    ("OFS sellers", "ofs_2_shares", "Shares 2"),
    ("OFS sellers", "ofs_2_cr", "Amt 2 Cr"),
    ("OFS sellers", "ofs_3_name", "Seller 3"),
    ("OFS sellers", "ofs_3_category", "Cat 3"),
    ("OFS sellers", "ofs_3_shares", "Shares 3"),
    ("OFS sellers", "ofs_3_cr", "Amt 3 Cr"),
    ("Company", "about_text", "About"),
    ("Company", "strengths", "Strengths"),
    ("Company", "company_address", "Address"),
    ("Company", "company_email", "Email"),
    ("Meta", "scraped_at", "Scraped at"),
    ("Meta", "parse_warnings", "Warnings"),
]

FIELD_KEYS = [f for _, f, _ in COLUMN_GROUPS]
GROUP_ROW = [g for g, _, _ in COLUMN_GROUPS]
LABEL_ROW = [lab for _, _, lab in COLUMN_GROUPS]

FIN_METRIC_MAP = {
    "assets": "assets",
    "total income": "total_income",
    "profit after tax": "pat",
    "ebitda": "ebitda",
    "net worth": "net_worth",
    "total borrowing": "borrowings",
}

_ILLEGAL_XLSX = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _xlsx_value(val: Any) -> Any:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, str):
        text = _ILLEGAL_XLSX.sub("", val)
        if text in {"", "None", "nan"}:
            return None
        return text
    return val


def _replace_or_fallback(tmp: Path, dest: Path) -> None:
    """Replace dest with tmp. If dest is locked, write dest.unlocked.* instead."""
    last_err: OSError | None = None
    for i in range(8):
        try:
            os.replace(tmp, dest)
            return
        except PermissionError as exc:
            last_err = exc
            time.sleep(0.4 * (i + 1))
    fallback = dest.with_name(f"{dest.stem}.unlocked{dest.suffix}")
    try:
        os.replace(tmp, fallback)
    except OSError:
        fallback.write_bytes(tmp.read_bytes())
        tmp.unlink(missing_ok=True)
    print(f"warning: {dest.name} is locked (close Excel/OneDrive); wrote {fallback.name} instead")
    if last_err and not fallback.exists():
        raise last_err


def _safe_to_csv(df: pd.DataFrame, path: Path, attempts: int = 8, header: bool = True) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    last_err: OSError | None = None
    for i in range(attempts):
        try:
            df.to_csv(path, index=False, encoding="utf-8-sig", header=header)
            return
        except PermissionError as exc:
            last_err = exc
            time.sleep(0.4 * (i + 1))
    fallback = path.with_name(f"{path.stem}.unlocked{path.suffix}")
    df.to_csv(fallback, index=False, encoding="utf-8-sig", header=header)
    print(f"warning: {path.name} is locked (close Excel/OneDrive sync); wrote {fallback.name} instead")
    if last_err and not fallback.exists():
        raise last_err


def _exch_prefix(name: str) -> Optional[str]:
    t = (name or "").lower()
    if "nse" in t:
        return "listing_nse"
    if "bse" in t:
        return "listing_bse"
    return None


def flatten_into_master(master: dict[str, Any], sats: dict[str, list[dict[str, Any]]] | None) -> dict[str, Any]:
    """Copy satellite tables onto the IPO row so one sheet holds everything."""
    row = dict(master)
    sats = sats or {}

    # Mirror latest financials onto fy1_* if the grid parse filled the short aliases.
    if row.get("fy1_period") is None:
        row["fy1_period"] = row.get("fy_latest")
        row["fy1_assets"] = row.get("assets_cr")
        row["fy1_total_income"] = row.get("total_income_cr")
        row["fy1_pat"] = row.get("pat_cr")
        row["fy1_ebitda"] = row.get("ebitda_cr")
        row["fy1_net_worth"] = row.get("net_worth_cr")
        row["fy1_borrowings"] = row.get("borrowings_cr")

    by_period: dict[str, dict[str, Any]] = {}
    for rec in sats.get("financials") or []:
        period = str(rec.get("period") or "")
        key = rec.get("metric_key") or ""
        dest = None
        for needle, field in FIN_METRIC_MAP.items():
            if needle in key:
                dest = field
                break
        if not period or not dest:
            continue
        by_period.setdefault(period, {})[dest] = rec.get("value")
    periods = sorted(by_period.keys(), reverse=True)[:3]
    for i, period in enumerate(periods, start=1):
        vals = by_period[period]
        row[f"fy{i}_period"] = period
        for field in ("assets", "total_income", "pat", "ebitda", "net_worth", "borrowings"):
            row[f"fy{i}_{field}"] = vals.get(field)

    for rec in sats.get("listing_day") or []:
        prefix = _exch_prefix(str(rec.get("exchange") or ""))
        if not prefix:
            continue
        row[f"{prefix}_open"] = rec.get("open")
        row[f"{prefix}_high"] = rec.get("high")
        row[f"{prefix}_low"] = rec.get("low")
        row[f"{prefix}_last"] = rec.get("last")

    for i, rec in enumerate((sats.get("objects") or [])[:4], start=1):
        row[f"object_{i}"] = rec.get("object")
        row[f"object_{i}_cr"] = rec.get("amount_cr")

    for i, rec in enumerate((sats.get("ofs_shareholders") or [])[:3], start=1):
        row[f"ofs_{i}_name"] = rec.get("name")
        row[f"ofs_{i}_category"] = rec.get("category")
        row[f"ofs_{i}_shares"] = rec.get("shares")
        row[f"ofs_{i}_cr"] = rec.get("amount_cr")

    history = sats.get("gmp_history") or []
    row["gmp_obs_count"] = len(history) if history else None
    return row


def _ordered_frame(rows: Iterable[dict[str, Any]]) -> pd.DataFrame:
    df = pd.DataFrame(list(rows))
    if df.empty:
        return pd.DataFrame(columns=FIELD_KEYS)
    missing = [c for c in FIELD_KEYS if c not in df.columns]
    if missing:
        df = pd.concat([df, pd.DataFrame({c: [None] * len(df) for c in missing})], axis=1)
    return df[FIELD_KEYS]


def _has_group_header(path: Path) -> bool:
    if not path.exists() or path.stat().st_size == 0:
        return False
    first = path.read_text(encoding="utf-8-sig", errors="replace").splitlines()[:1]
    return bool(first) and first[0].lstrip("\ufeff").startswith("Identity")


def read_master(path: Path) -> pd.DataFrame:
    if not path.exists() or path.stat().st_size == 0:
        return pd.DataFrame(columns=FIELD_KEYS)
    # Subcolumn labels repeat (Period, Open, Last, …). Bind columns by position.
    if _has_group_header(path):
        df = pd.read_csv(path, header=None, skiprows=2, dtype=str, keep_default_na=False)
        n = len(FIELD_KEYS)
        if df.empty:
            return pd.DataFrame(columns=FIELD_KEYS)
        if df.shape[1] < n:
            for i in range(df.shape[1], n):
                df[i] = ""
        df = df.iloc[:, :n]
        df.columns = FIELD_KEYS
        return df
    df = pd.read_csv(path, header=0, dtype=str, keep_default_na=False)
    for col in FIELD_KEYS:
        if col not in df.columns:
            df[col] = ""
    return df[FIELD_KEYS] if all(c in df.columns for c in FIELD_KEYS) else df


def write_master_csv(path: Path, rows: Iterable[dict[str, Any]]) -> None:
    data = _ordered_frame(rows)
    header = pd.DataFrame([dict(zip(FIELD_KEYS, LABEL_ROW))], columns=FIELD_KEYS)
    out = pd.concat([header, data], ignore_index=True)
    path.parent.mkdir(parents=True, exist_ok=True)
    group_line = ",".join(_csv_escape(g) for g in GROUP_ROW)
    body = out.to_csv(index=False, header=False)
    text = group_line + "\n" + body
    tmp = path.with_name(path.name + ".tmp")
    tmp.write_text("\ufeff" + text, encoding="utf-8")
    try:
        _replace_or_fallback(tmp, path)
    finally:
        tmp.unlink(missing_ok=True)


def _csv_escape(value: str) -> str:
    if any(ch in value for ch in ',\"\n'):
        return '"' + value.replace('"', '""') + '"'
    return value


def write_master_xlsx(path: Path, rows: Iterable[dict[str, Any]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("warning: openpyxl not installed; skipping ipos.xlsx")
        return
    data = _ordered_frame(rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "IPOs"
    fills = [
        PatternFill("solid", fgColor="1F4E79"),
        PatternFill("solid", fgColor="2E75B6"),
        PatternFill("solid", fgColor="5B9BD5"),
    ]
    header_font = Font(bold=True, color="FFFFFF", size=10)
    sub_font = Font(bold=True, size=9)
    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    # Group row + merge
    col = 1
    fill_i = 0
    while col <= len(COLUMN_GROUPS):
        group = COLUMN_GROUPS[col - 1][0]
        start = col
        while col <= len(COLUMN_GROUPS) and COLUMN_GROUPS[col - 1][0] == group:
            col += 1
        end = col - 1
        fill = fills[fill_i % len(fills)]
        fill_i += 1
        cell = ws.cell(1, start, group)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for c in range(start, end + 1):
            ws.cell(1, c).fill = fill
            ws.cell(1, c).font = header_font
            ws.cell(1, c).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(1, c).border = thin
        if end > start:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
    for i, label in enumerate(LABEL_ROW, start=1):
        cell = ws.cell(2, i, label)
        cell.font = sub_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = thin
    for r_i, rec in enumerate(data.to_dict(orient="records"), start=3):
        for c_i, key in enumerate(FIELD_KEYS, start=1):
            val = rec.get(key)
            if val is None or (isinstance(val, float) and pd.isna(val)):
                val = None
            val = _xlsx_value(val)
            cell = ws.cell(r_i, c_i, val)
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=False)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(FIELD_KEYS))}{max(2, 2 + len(data))}"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 32
    for i, key in enumerate(FIELD_KEYS, start=1):
        width = 14
        if key in {"url", "about_text", "strengths", "company_address", "promoters", "lead_managers"}:
            width = 28
        if key.startswith("object_"):
            width = 24
        ws.column_dimensions[get_column_letter(i)].width = width
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(path.name + ".tmp")
    try:
        wb.save(tmp)
        _replace_or_fallback(tmp, path)
    except PermissionError:
        fallback = path.with_name(f"{path.stem}.unlocked{path.suffix}")
        wb.save(fallback)
        print(f"warning: {path.name} is locked; wrote {fallback.name} instead")
    finally:
        tmp.unlink(missing_ok=True)


def append_master(
    out_dir: Path,
    rows: list[dict[str, Any]],
    key: str = "ipo_id",
    write_xlsx: bool = False,
) -> None:
    csv_path = out_dir / "ipos.csv"
    new = _ordered_frame(rows)
    if csv_path.exists() and csv_path.stat().st_size > 0:
        old = read_master(csv_path)
        ids = set(new[key].astype(str))
        if key in old.columns:
            old = old[~old[key].astype(str).isin(ids)]
        combined = pd.concat([old, new.astype(str)], ignore_index=True)
    else:
        combined = new
    records = combined.to_dict(orient="records")
    write_master_csv(csv_path, records)
    if write_xlsx:
        write_master_xlsx(out_dir / "ipos.xlsx", records)


def rebuild_master_xlsx(out_dir: Path) -> None:
    """Rebuild ipos.xlsx from ipos.csv (CSV is the source of truth)."""
    csv_path = out_dir / "ipos.csv"
    if not csv_path.exists() or csv_path.stat().st_size == 0:
        print("no ipos.csv to rebuild from")
        return
    rows = read_master(csv_path).to_dict(orient="records")
    write_master_xlsx(out_dir / "ipos.xlsx", rows)
    print(f"wrote ipos.xlsx ({len(rows)} rows)", flush=True)


LEGACY_OUTPUTS = (
    "financials.csv",
    "kpis.csv",
    "listing_day.csv",
    "objects.csv",
    "ofs_shareholders.csv",
    "reservation.csv",
    "subscription.csv",
    "gmp_history.csv",
    "ipo_index.csv",
    "financials.unlocked.csv",
    "kpis.unlocked.csv",
    "listing_day.unlocked.csv",
    "ipos.csv.tmp",
    "ipos.xlsx.tmp",
)


def remove_legacy_outputs(out_dir: Path) -> None:
    """Delete leftover satellite files from the old multi-CSV layout."""
    if not out_dir.exists():
        return
    for name in LEGACY_OUTPUTS:
        path = out_dir / name
        if path.exists():
            path.unlink()
            print(f"removed leftover {name}")
    for tmp in out_dir.glob("*.tmp"):
        tmp.unlink()
        print(f"removed leftover {tmp.name}")


def write_csv(path: Path, rows: Iterable[dict[str, Any]], columns: list[str] | None = None) -> None:
    df = pd.DataFrame(list(rows))
    if columns:
        for col in columns:
            if col not in df.columns:
                df[col] = None
        extra = [c for c in df.columns if c not in columns]
        df = df[list(columns) + extra]
    _safe_to_csv(df, path)


def append_or_replace(path: Path, rows: list[dict[str, Any]], key: str = "ipo_id") -> None:
    new = pd.DataFrame(rows)
    if new.empty:
        if not path.exists():
            _safe_to_csv(new, path)
        return
    if path.exists() and path.stat().st_size > 0:
        old = pd.read_csv(path, dtype=str, keep_default_na=False)
        ids = set(new[key].astype(str))
        if key in old.columns:
            old = old[~old[key].astype(str).isin(ids)]
        df = pd.concat([old, new.astype(str)], ignore_index=True)
    else:
        df = new
    _safe_to_csv(df, path)


def load_done_ids(path: Path) -> set[str]:
    if not path.exists() or path.stat().st_size == 0:
        return set()
    df = read_master(path) if path.name == "ipos.csv" or _has_group_header(path) else pd.read_csv(
        path, dtype=str, keep_default_na=False
    )
    if "ipo_id" not in df.columns:
        return set()
    return {str(v) for v in df["ipo_id"] if str(v) not in {"", "nan", "None", "IPO ID"}}


def coverage_report(masters: list[dict[str, Any]]) -> str:
    n = len(masters) or 1

    def pct(field: str) -> str:
        c = sum(1 for m in masters if m.get(field) not in (None, "", float("nan")))
        return f"{c}/{len(masters)} ({100 * c / n:.1f}%)"

    lines = [
        f"IPOs in master: {len(masters)}",
        f"  subscription: {pct('sub_total_x')}",
        f"  financials (PAT): {pct('fy1_pat')}",
        f"  ISIN: {pct('isin')}",
        f"  listing OHLC: {pct('listing_nse_open')}",
        f"  GMP close: {pct('gmp_rs')}",
        f"  industry: {pct('industry')}",
        f"  ROCE: {pct('roce')}",
        "  CFO/FCF: not published on Chittorgarh IPO pages (skipped)",
    ]
    return "\n".join(lines)
