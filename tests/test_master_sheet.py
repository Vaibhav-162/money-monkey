from __future__ import annotations

from pathlib import Path

from chittorgarh.export import COLUMN_GROUPS, flatten_into_master, read_master, write_master_csv, write_master_xlsx
from chittorgarh.parse_ipo import parse_ipo_html

FIXTURE = Path(__file__).resolve().parent / "fixtures" / "lohia_2574.html"
URL = "https://www.chittorgarh.com/ipo/lohia-corp-ipo/2574/"


def test_master_sheet_two_header_rows(tmp_path: Path) -> None:
    html = FIXTURE.read_text(encoding="utf-8", errors="replace")
    parsed = parse_ipo_html(html, URL, "mainboard", 2026, {"ipo_id": "2574"})
    row = flatten_into_master(parsed["master"], parsed["satellites"])
    csv_path = tmp_path / "ipos.csv"
    xlsx_path = tmp_path / "ipos.xlsx"
    write_master_csv(csv_path, [row])
    write_master_xlsx(xlsx_path, [row])

    lines = csv_path.read_text(encoding="utf-8-sig").splitlines()
    assert lines[0].startswith("Identity")
    assert "Classification" in lines[0]
    assert "Financials FY1 (latest)" in lines[0]
    assert "Listing day NSE" in lines[0]
    df = read_master(csv_path)
    assert len(df) == 1
    assert str(df.iloc[0]["ipo_id"]) == "2574"
    assert "Industrial" in str(df.iloc[0]["industry"])
    assert float(df.iloc[0]["listing_nse_open"]) == 461
    assert xlsx_path.exists()

    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws = wb.active
    assert ws.cell(1, 1).value == "Identity"
    identity_width = sum(1 for g, _, _ in COLUMN_GROUPS if g == "Identity")
    merged = next(
        r for r in ws.merged_cells.ranges if r.min_row == 1 and r.min_col == 1
    )
    assert merged.max_col == identity_width
    assert ws.cell(2, 1).value == "IPO ID"
    assert str(ws.cell(3, 1).value) == "2574"


def test_remove_legacy_outputs(tmp_path: Path) -> None:
    from chittorgarh.export import remove_legacy_outputs

    (tmp_path / "financials.csv").write_text("x", encoding="utf-8")
    (tmp_path / "ipo_index.csv").write_text("x", encoding="utf-8")
    (tmp_path / "ipos.csv").write_text("keep", encoding="utf-8")
    remove_legacy_outputs(tmp_path)
    assert not (tmp_path / "financials.csv").exists()
    assert not (tmp_path / "ipo_index.csv").exists()
    assert (tmp_path / "ipos.csv").read_text(encoding="utf-8") == "keep"


def test_append_master_keeps_one_file(tmp_path: Path) -> None:
    from chittorgarh.export import append_master, read_master

    append_master(tmp_path, [{"ipo_id": "1", "company_name": "Alpha", "industry": "Old"}])
    append_master(tmp_path, [{"ipo_id": "2", "company_name": "Beta"}])
    append_master(tmp_path, [{"ipo_id": "1", "company_name": "Alpha", "industry": "Updated"}])
    df = read_master(tmp_path / "ipos.csv")
    assert len(df) == 2
    one = df[df["ipo_id"].astype(str) == "1"].iloc[0]
    assert str(one["industry"]) == "Updated"
    assert set(df["ipo_id"].astype(str)) == {"1", "2"}
    assert not (tmp_path / "ipos.xlsx").exists()

    from chittorgarh.export import rebuild_master_xlsx
    rebuild_master_xlsx(tmp_path)
    assert (tmp_path / "ipos.xlsx").exists()
    from openpyxl import load_workbook
    wb = load_workbook(tmp_path / "ipos.xlsx")
    assert str(wb.active.cell(3, 1).value) in {"1", "2"}
